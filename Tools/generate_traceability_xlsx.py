"""
Script to generate TRACEABILITY_MATRIX.xlsx with openpyxl.
Contains multiple filterable sheets per Epic and Compliance mapping, formatted professionally.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TARGET_XLSX = r'd:\justificadll\MarkDown\TRACEABILITY_MATRIX.xlsx'

wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
data_font = Font(name="Segoe UI", size=10, color="000000")
bold_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
border_thin = Side(border_style="thin", color="D9D9D9")
border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Master Data
data_all = [
    ("UC-01", "Melakukan Registrasi Klien", "Klien", "A.UC-01", "AD-01", "SD-01, SD-08", "Auth (#mockup_auth)", "ST-001", "UU PDP Art 15,16 / UU 17/2023", "FieldEnc(NIK), Consent v1, Dukcapil Sync", "Epic 1 - Core & Auth"),
    ("UC-02", "Melakukan Login Klien", "Klien", "A.UC-02", "AD-02", "SD-02", "Auth (#mockup_auth)", "ST-002", "UU PDP Art 15,16 / Permenkes 24/2022", "MFA/OTP, Audit Trail WORM, JWT", "Epic 1 - Core & Auth"),
    ("UC-03", "Memilih Mitra Profesional", "Klien", "A.UC-03", "AD-03", "SD-03, SD-09", "Mitra (#mockup_mitra)", "ST-005", "UU 17/2023, UU 18/2003, HIMPSI", "Active License Check (STR/SIPP/Peradi), Geolocation", "Epic 1 - Core & Auth"),
    ("UC-04", "Melakukan Konsultasi Real-Time", "Klien, Mitra", "A.UC-04", "AD-04", "SD-04", "Chat Room (#mockup_chat_room)", "ST-008", "Permenkes 24/2022, UU 18/2003", "E2EE Chat, ZK Privilege, WORM Lock, Timer", "Epic 2 - Comm & Payment"),
    ("UC-05", "Melakukan Pembayaran", "Klien", "A.UC-05", "AD-05", "SD-05", "Payment (#mockup_payment_gateway)", "ST-007", "PCI-DSS, PBI, UU 18/2003", "Escrow Pro Bono, Webhook SHA-256, Snap Token", "Epic 2 - Comm & Payment"),
    ("UC-06", "Memberikan Ulasan dan Rating", "Klien", "A.UC-06", "AD-06", "SD-06, SD-10", "Feedback (#mockup_feedback)", "ST-021", "Permenkes 17/2023, UU 18/2003", "Mandatory Adverse Event Form (Rating<=2), Anonimisasi Hukum", "Epic 4 - Admin & Feedback"),
    ("UC-07", "Registrasi Mitra Profesional", "Mitra", "B.UC-07", "AD-07", "SD-07, SD-11", "Mitra Auth (#mockup_auth)", "ST-003", "UU 17/2023, UU 18/2003, HIMPSI", "STR/SIP/SIPP/KTA Upload, WORM Storage, KKI Sync", "Epic 1 - Core & Auth"),
    ("UC-08", "Login Mitra Profesional", "Mitra", "B.UC-08", "AD-08", "SD-12", "Mitra Auth (#mockup_auth)", "ST-004", "Permenkes 24/2022, UU PDP", "Mandatory MFA TOTP, Audit Trail WORM, Role Auth", "Epic 1 - Core & Auth"),
    ("UC-09", "Konfirmasi Ketersediaan (On/Off)", "Mitra", "B.UC-09", "AD-09", "SD-13", "Dashboard Mitra (#mockup_dashboard_mitra)", "ST-006", "SIRS RS Sync, SIPP Pengadilan Sync", "Real-time Faskes/Sidang Sync, 30m Psychology Buffer", "Epic 1 - Core & Auth"),
    ("UC-10", "Melayani Konsultasi", "Mitra", "B.UC-10", "AD-04", "SD-04", "Chat Room (#mockup_chat_room)", "ST-009", "Permenkes 24/2022 (SLA 5m)", "5-min SLA Alert, Auto-refund timeout, E2EE", "Epic 2 - Comm & Payment"),
    ("UC-11", "Membuat Catatan Sesi Konsultasi", "Mitra", "B.UC-11", "AD-10", "SD-14", "Chat Room Panel (#mockup_chat_room)", "ST-011", "Permenkes 24/2022, HIMPSI, UU 18/2003", "SOAP+ICD-10, DAP+Crisis Flag, IRAC Case Memo, WORM 10Y", "Epic 3 - Domain Specific"),
    ("UC-12", "Menerbitkan Output Dokumen", "Mitra", "B.UC-12", "AD-11", "SD-15", "Chat Room Panel (#mockup_chat_room)", "ST-012", "Permenkes 73/2016, UU 10/2020", "e-Resep DDI Checker, e-Meterai Peruri Rp 10k, Controlled Drugs", "Epic 3 - Domain Specific"),
    ("UC-13", "Verifikasi Kredensial & SKTM Pro Bono", "Admin", "C.UC-13", "AD-12", "SD-16", "Admin Verifikasi (#mockup_admin_verif)", "ST-022", "UU 17/2023, HIMPSI, UU 18/2003", "Cross-check API KKI/HIMPSI/Peradi, Dukcapil & DTKS Sync", "Epic 4 - Admin & Feedback"),
    ("UC-14", "Mengelola Akun Klien (Suspend)", "Admin", "C.UC-14", "AD-13", "SD-17", "Admin Akun (#mockup_admin_akun)", "ST-023", "Due Process, UU PDP Art 15", "Warning 1/2/3, WORM Evidence, 14-Day Appeal Window", "Epic 4 - Admin & Feedback"),
    ("UC-15", "Mengelola Akun Mitra (Ethics Flow)", "Admin", "C.UC-15", "AD-14", "SD-18", "Admin Kasus Etik (#mockup_admin_etik)", "ST-024", "UU 17/2023, HIMPSI, Peradi", "Ethics Committee Hearing (4 Ahli), WORM Audit, National Body Report", "Epic 4 - Admin & Feedback"),
    ("UC-16", "Memantau Laporan Transaksi", "Admin", "C.UC-16", "AD-15", "SD-19", "Admin Keuangan (#mockup_admin_finance)", "ST-025", "SAK, Revenue Share Policy", "Proporsional Share (Kes 15%, Psi 20%, Huk 25%), WORM SHA-256", "Epic 4 - Admin & Feedback"),
    ("UC-17", "Mengelola Saldo & Penarikan Dana", "Mitra", "B.UC-17", "AD-16", "SD-20", "Dashboard Mitra (#mockup_dashboard_mitra)", "ST-026", "PPh 21 Dirjen Pajak, AML Faskes", "Threshold Control (<5M Auto, >=5M Manual Approval), PPh 21 Auto", "Epic 4 - Admin & Feedback"),
    ("Kes-UC01", "Menebus Resep & Membeli Obat", "Klien", "D.Kes-UC01", "AD-11", "SD-15", "Apotek Module (#mockup_modul_medis)", "ST-013", "Permenkes 73/2016", "SIA Apotek Sync, DDI Checker, Controlled Drug 3-Rangkap", "Epic 3 - Domain Specific"),
    ("Kes-UC02", "Membuat Janji Temu RS Offline", "Klien", "D.Kes-UC02", "AD-03", "SD-09", "RS Offline Booking (#mockup_modul_medis)", "ST-014", "API SIRS Rumah Sakit", "SIRS Quota Sync, BPJS Rujukan Validation", "Epic 3 - Domain Specific"),
    ("Kes-UC03", "Melihat Rekam Medis & Family Care", "Klien", "D.Kes-UC03", "AD-10", "SD-14", "Rekam Medis (#mockup_modul_medis)", "ST-015", "Permenkes 24/2022, UU PDP Art 16", "PIN/MFA Access, Digital Guardianship Consent, Read-Only WORM", "Epic 3 - Domain Specific"),
    ("Psi-UC01", "Mengisi Jurnal Mood Harian", "Klien", "D.Psi-UC01", "AD-17", "SD-21", "Mood Tracker (#mockup_modul_psikologi)", "ST-016", "Kode Etik HIMPSI Bab IV", "Field-level Enc, Read-Only by Active Counselor, Proactive Wellness Alert", "Epic 3 - Domain Specific"),
    ("Psi-UC02", "Mengakses Audio Meditasi", "Klien", "D.Psi-UC02", "AD-18", "SD-22", "Relaxation (#mockup_modul_psikologi)", "ST-017", "HIMPSI Psychoeducation", "Curated Mindfulness Audio, Adaptive Bitrate Streaming", "Epic 3 - Domain Specific"),
    ("Psi-UC03", "Mengisi Tes Asesmen DASS-21", "Klien", "D.Psi-UC03", "AD-19", "SD-23", "DASS-21 (#mockup_modul_psikologi)", "ST-018", "HIMPSI Mandatory Crisis Protocol", "Scoring Engine, Mandatory Crisis Protocol (Hotline 119 ext 8, 10s Lock)", "Epic 3 - Domain Specific"),
    ("Huk-UC01", "Mengunggah Berkas Perkara", "Klien", "D.Huk-UC01", "AD-04", "SD-04", "Chat Room (#mockup_chat_room)", "ST-010", "UU 18/2003 Art 19 (Privilege)", "Zero-Knowledge E2EE, Privileged Evidence Stamp, Malware Scan", "Epic 2 - Comm & Payment"),
    ("Huk-UC02", "Membuat Draf Dokumen Hukum", "Mitra", "D.Huk-UC02", "AD-11", "SD-15", "Legal Drafting (#mockup_modul_hukum)", "ST-019", "UU 18/2003, UU 10/2020 e-Meterai", "IRAC Template Engine, e-Meterai Peruri Rp 10k, Version Control v1/v2/Final", "Epic 3 - Domain Specific"),
    ("Huk-UC03", "Melakukan Konsultasi Pro Bono", "Klien", "D.Huk-UC03", "AD-12", "SD-16", "Pro Bono (#mockup_modul_hukum)", "ST-020", "UU 18/2003 Art 22 Pro Bono", "SKTM Dukcapil/DTKS Sync, Escrow Subsidi Rp 0, Quota Max 3/Mo", "Epic 3 - Domain Specific")
]

headers = ["UC-ID", "Use Case Name", "Actor Utama", "Scenario Doc", "Activity Diagram", "Sequence Diagram", "Mockup Page", "Backlog Story", "Regulasi Utama", "Compliance Flags", "Epic / Sprint"]

def create_sheet(sheet_name, data_rows):
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # Title Row
    ws.merge_cells("A1:K1")
    ws["A1"] = f"JUSTIFICA 3-IN-1 — TRACEABILITY MATRIX: {sheet_name.upper()}"
    ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Header Row
    ws.append([]) # Row 2 blank
    ws.append(headers) # Row 3
    ws.row_dimensions[3].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_box
        
    # Data Rows
    for row_idx, row_data in enumerate(data_rows, start=4):
        ws.append(list(row_data))
        ws.row_dimensions[row_idx].height = 28
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_box
            if col_idx in [1, 3, 5, 6, 8]:
                cell.alignment = align_center
                if col_idx in [1, 8]:
                    cell.font = bold_font
            else:
                cell.alignment = align_left
                
            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Enable AutoFilter on row 3
    ws.auto_filter.ref = f"A3:K{len(data_rows)+3}"
    return ws

# 1. Master Sheet
create_sheet("All Use Cases (26)", data_all)

# 2. Epic Sheets
epics = [
    ("Epic 1 - Core & Auth", [r for r in data_all if "Epic 1" in r[10]]),
    ("Epic 2 - Comm & Payment", [r for r in data_all if "Epic 2" in r[10]]),
    ("Epic 3 - Domain Specific", [r for r in data_all if "Epic 3" in r[10]]),
    ("Epic 4 - Admin & Feedback", [r for r in data_all if "Epic 4" in r[10]])
]

for epic_name, epic_rows in epics:
    create_sheet(epic_name[:31], epic_rows)

# 3. Compliance Summary Sheet
ws_comp = wb.create_sheet(title="Compliance Mapping")
ws_comp.views.sheetView[0].showGridLines = True
ws_comp.merge_cells("A1:F1")
ws_comp["A1"] = "JUSTIFICA 3-IN-1 — REGULATORY COMPLIANCE MAPPING MATRIX"
ws_comp["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
ws_comp.row_dimensions[1].height = 30

comp_headers = ["Regulasi Utama", "Pasal / Bab", "Domain Relevan", "Fitur / Modul Utama", "Compliance Mechanism & Guardrail", "Terkait UC-ID"]
ws_comp.append([])
ws_comp.append(comp_headers)
ws_comp.row_dimensions[3].height = 25
for col_idx in range(1, len(comp_headers) + 1):
    cell = ws_comp.cell(row=3, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center
    cell.border = border_box

comp_data = [
    ("UU No. 17 Tahun 2023 (Kesehatan)", "Pasal 172, 173, 184", "Kesehatan Medis", "Registrasi Dokter, e-Resep, Telemedicine", "Cross-check STR/SIP via API KKI/KTKI; validasi aktif; DDI Checker wajib; e-Resep digital tersertifikasi.", "UC-03, UC-07, UC-12, Kes-UC01"),
    ("Permenkes No. 24 Tahun 2022", "Pasal 4, 11, 15", "Kesehatan Medis", "Rekam Medis Elektronik (EME)", "Enkripsi E2EE saat transmisi; field-level encryption di DB; MFA wajib bagi dokter/pasien; SOAP Note & ICD-10.", "UC-02, UC-04, UC-08, UC-11, Kes-UC03"),
    ("Permenkes No. 73 Tahun 2016", "Pasal 3, 5, 6", "Kesehatan Medis", "Standar Pelayanan Farmasi & e-Resep", "Integrasi SIA Apotek Mitra; telaah resep apoteker; Controlled Drugs 3 rangkap digital + verifikasi KTP fisik.", "UC-12, Kes-UC01"),
    ("Kode Etik HIMPSI (2019)", "Bab III, IV, V", "Psikologi", "Registrasi Psikolog, Konseling, Asesmen", "SIPP validasi; mandatory buffer 30 menit antar sesi; DAP Note; kerahasiaan mood tracker & catatan konseling.", "UC-03, UC-07, UC-09, UC-11, Psi-UC01"),
    ("HIMPSI Mandatory Crisis Protocol", "Prosedur Intervensi Krisis", "Psikologi", "Asesmen DASS-21 & Konseling", "Jika skor DASS-21 Severe/Extreme: pop-up wajib Hotline 119 ext 8 (10s read lock); alert real-time ke Supervisor Klinis.", "UC-11, Psi-UC03"),
    ("UU No. 18 Tahun 2003 (Advokat)", "Pasal 18, 19, 20", "Hukum", "Advocate-Client Privilege & Kerahasiaan", "Zero-Knowledge E2EE Chat (Admin dilarang/tidak bisa baca); penandaan PRIVILEGED AND CONFIDENTIAL; anonimisasi ulasan.", "UC-04, UC-06, Huk-UC01, Huk-UC02"),
    ("UU No. 18 Tahun 2003 Pasal 22", "Pasal 22 (Bantuan Hukum)", "Hukum Pro Bono", "Layanan Hukum Cuma-Cuma (Pro Bono)", "Validasi SKTM via cross-check API Dukcapil & DTKS Kemensos; subsidi Rp 0; sistem escrow platform; max 3 kasus/bln.", "UC-13, Huk-UC03"),
    ("UU No. 10 Tahun 2020", "Pasal 3, 5 (Bea Meterai)", "Hukum", "Legal Drafting & Akta Hukum", "Pembubuhan e-Meterai resmi Rp 10.000 via integrasi API Perum Peruri; version control v1/v2/Final; retensi WORM 10 thn.", "UC-12, Huk-UC02"),
    ("UU PDP No. 27 Tahun 2022", "Pasal 15, 16, 17, 26", "Shared All Domain", "Pelindungan Data Pribadi & Family Care", "Granular consent terpisah per domain; Digital Guardianship Consent untuk dewasa; WORM audit trail logging abadi.", "UC-01, UC-02, UC-08, Kes-UC03"),
    ("Due Process of Law & SAK", "Prinsip Perlindungan Konsumen", "Admin & Keuangan", "Manajemen Akun & Laporan Keuangan", "Warning bertingkat 1/2/3; masa banding 14 hari kerja; Tim Etik Multidisiplin 4 ahli; ekspor WORM SHA-256 hash.", "UC-14, UC-15, UC-16, UC-17")
]

for row_idx, row_data in enumerate(comp_data, start=4):
    ws_comp.append(list(row_data))
    ws_comp.row_dimensions[row_idx].height = 32
    for col_idx in range(1, len(row_data) + 1):
        cell = ws_comp.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = border_box
        if col_idx in [1, 2, 3, 6]:
            cell.alignment = align_center
            if col_idx == 1:
                cell.font = bold_font
        else:
            cell.alignment = align_left
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

for col in ws_comp.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_comp.column_dimensions[col_letter].width = max(max_len + 3, 14)
ws_comp.auto_filter.ref = f"A3:F{len(comp_data)+3}"

wb.save(TARGET_XLSX)
print(f"Successfully generated {TARGET_XLSX}")
